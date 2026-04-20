from io import BytesIO
import logging
from datetime import datetime

from django.contrib import messages
from django.contrib.auth import login, logout
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST
from django.utils.timezone import localtime, now
from django.utils.translation import get_language, activate, gettext as _
from django.middleware.locale import LocaleMiddleware
from reportlab.pdfgen import canvas

from .forms import AbsenceJustificationForm, CustomAuthenticationForm, EtudiantForm, FiliereForm, SignUpForm, ImportExcelForm, PasswordChangeForm, UserProfileUpdateForm
from .models import AbsenceJustification, ClassSchedule, Etudiant, Filiere, Presence

# Configure logging
logger = logging.getLogger(__name__)

WEEKDAY_ORDER = {
    'Monday': 0,
    'Tuesday': 1,
    'Wednesday': 2,
    'Thursday': 3,
    'Friday': 4,
    'Saturday': 5,
    'Sunday': 6,
}

# Importer openpyxl pour lire les fichiers Excel
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Alternative: utiliser xlrd pour les anciens fichiers Excel
try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False


# ======================== LANGUAGE DETECTION ========================

def detect_language_and_redirect(request, target_url='login'):
    """
    Detect user's preferred language and redirect to language-prefixed URL.
    Checks in order: session, browser Accept-Language, default language
    """
    # Get language from session (if user already selected)
    language = request.session.get('django_language')
    
    # If not in session, get from browser Accept-Language header
    if not language:
        accept_language = request.META.get('HTTP_ACCEPT_LANGUAGE', '')
        # Parse first preference
        if accept_language:
            lang = accept_language.split(',')[0].split('-')[0].lower()
            # Check if it's a supported language
            supported_langs = ['en', 'fr', 'ar']
            if lang in supported_langs:
                language = lang
    
    # Default to French if no language detected
    if not language:
        language = 'fr'
    
    # If user is logged in and trying to access login, redirect to dashboard
    if target_url == 'login' and request.user.is_authenticated:
        return redirect(f'/{language}/dashboard/')
    
    # Redirect to language-prefixed URL
    return redirect(f'/{language}/{target_url}/')


def root_language_redirect(request):
    """
    Redirect root URL to language-prefixed home page.
    Detects language from session, browser, or defaults to French.
    """
    if request.user.is_authenticated:
        return detect_language_and_redirect(request, 'dashboard')
    return detect_language_and_redirect(request, 'login')


# ======================== HELPER FUNCTIONS ========================

def get_student_for_user(user):
    """Safely get student profile for a user."""
    try:
        if not user.is_authenticated:
            return None
        return getattr(user, 'etudiant_profile', None)
    except Exception as e:
        logger.error(f"Error getting student profile for user {user.username}: {str(e)}")
        return None


def redirect_after_login(user):
    """Redirect user to appropriate page after login."""
    try:
        student = get_student_for_user(user)
        if student:
            return redirect('student_dashboard')
        return redirect('dashboard')
    except Exception as e:
        logger.error(f"Error redirecting user {user.username}: {str(e)}")
        return redirect('dashboard')


def validate_file_size(file, max_size_mb=5):
    """Validate uploaded file size."""
    if file.size > max_size_mb * 1024 * 1024:
        return False, _("File is too large. Maximum size is {max_size_mb}MB").format(max_size_mb=max_size_mb)
    return True, _("File size is valid")


def validate_excel_file(file):
    """Validate Excel file structure."""
    try:
        if not file.name.endswith(('.xlsx', '.xls')):
            return False, _("Invalid file format. Please upload .xlsx or .xls file")
        return True, _("File format is valid")
    except Exception as e:
        logger.error(f"Error validating Excel file: {str(e)}")
        return False, _("Error validating file")


def get_attendance_stats(filiere):
    """Get attendance statistics for a filiere."""
    try:
        today = now().date()
        total_students = filiere.etudiants.count()
        present_today = Presence.objects.filter(
            etudiant__filiere=filiere,
            date=today,
            present=True
        ).count()
        absent_today = Presence.objects.filter(
            etudiant__filiere=filiere,
            date=today,
            present=False
        ).count()
        
        attendance_rate = 0
        if total_students > 0:
            attendance_rate = (present_today / total_students) * 100
            
        return {
            'total': total_students,
            'present': present_today,
            'absent': absent_today,
            'rate': round(attendance_rate, 2)
        }
    except Exception as e:
        logger.error(f"Error getting attendance stats for filiere {filiere.id}: {str(e)}")
        return {'total': 0, 'present': 0, 'absent': 0, 'rate': 0}


def log_user_action(user, action, details=""):
    """Log user actions for audit trail."""
    try:
        log_message = f"User {user.username} - Action: {action}"
        if details:
            log_message += f" - Details: {details}"
        logger.info(log_message)
    except Exception as e:
        logger.error(f"Error logging user action: {str(e)}")


def build_schedule_section(filiere, schedules, current_day_name, current_time):
    """Build an ordered weekly calendar summary for a filiere."""
    ordered_schedules = sorted(
        schedules,
        key=lambda schedule: (
            WEEKDAY_ORDER.get(schedule.day_of_week, 99),
            schedule.start_time,
            schedule.end_time,
        ),
    )

    grouped_days = []
    day_lookup = {}
    next_session = None
    today_session_count = 0

    for schedule in ordered_schedules:
        if schedule.day_of_week == current_day_name:
            today_session_count += 1

        if next_session is None:
            is_future_day = WEEKDAY_ORDER.get(schedule.day_of_week, 99) > WEEKDAY_ORDER.get(current_day_name, 99)
            is_later_today = schedule.day_of_week == current_day_name and schedule.start_time >= current_time
            if is_future_day or is_later_today:
                next_session = schedule

        label = schedule.get_day_of_week_display()
        if label not in day_lookup:
            day_lookup[label] = {
                'label': label,
                'count': 0,
                'items': [],
            }
            grouped_days.append(day_lookup[label])
        day_lookup[label]['items'].append(schedule)
        day_lookup[label]['count'] += 1

    first_session = ordered_schedules[0] if ordered_schedules else None
    last_session = ordered_schedules[-1] if ordered_schedules else None

    return {
        'filiere': filiere,
        'grouped_days': grouped_days,
        'total_sessions': len(ordered_schedules),
        'occupied_days': len(grouped_days),
        'today_session_count': today_session_count,
        'first_session': first_session,
        'last_session': last_session,
        'next_session': next_session,
    }


# ======================== VIEWS ========================

def home(request):
    """Home page - redirect to appropriate dashboard."""
    if request.user.is_authenticated:
        return redirect_after_login(request.user)
    return redirect('login')


def login_view(request):
    if request.user.is_authenticated:
        return redirect_after_login(request.user)

    form = CustomAuthenticationForm(request, data=request.POST or None)
    if request.method == 'POST' and form.is_valid():
        login(request, form.get_user())
        return redirect_after_login(form.get_user())

    return render(request, 'login.html', {'form': form})


def signup_view(request):
    if request.user.is_authenticated:
        return redirect_after_login(request.user)

    form = SignUpForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        user = form.save()
        login(request, user)
        messages.success(request, _("Votre compte a ete cree avec succes."))
        return redirect('dashboard')

    return render(request, 'signup.html', {'form': form})


@login_required
def logout_view(request):
    logout(request)
    log_user_action(request.user, "LOGOUT")
    return redirect('login')


@login_required
def user_profile(request):
    """Display user profile page."""
    try:
        context = {
            'user': request.user,
            'student': get_student_for_user(request.user),
        }
        log_user_action(request.user, "VIEW_PROFILE")
        return render(request, 'user_profile.html', context)
    except Exception as e:
        logger.error(f"Error loading user profile for {request.user.username}: {str(e)}")
        messages.error(request, _("Error loading profile page"))
        return redirect('dashboard')


@login_required
def change_password(request):
    """Change user password."""
    try:
        if request.method == 'POST':
            form = PasswordChangeForm(request.user, request.POST)
            if form.is_valid():
                user = request.user
                user.set_password(form.cleaned_data['new_password1'])
                user.save()
                messages.success(request, _("Votre mot de passe a été changé avec succès."))
                log_user_action(user, "CHANGE_PASSWORD")
                logout(request)
                return redirect('login')
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f"{field}: {error}")
        else:
            form = PasswordChangeForm(request.user)
        
        return render(request, 'change_password.html', {'form': form})
    except Exception as e:
        logger.error(f"Error changing password for user {request.user.username}: {str(e)}")
        messages.error(request, _("Erreur lors du changement de mot de passe"))
        return redirect('dashboard')


@login_required
def update_profile(request):
    """Update user profile information."""
    try:
        if request.method == 'POST':
            form = UserProfileUpdateForm(request.POST, instance=request.user)
            if form.is_valid():
                form.save()
                messages.success(request, _("Votre profil a été mis à jour avec succès."))
                log_user_action(request.user, "UPDATE_PROFILE")
                return redirect('user_profile')
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f"{field}: {error}")
        else:
            form = UserProfileUpdateForm(instance=request.user)
        
        return render(request, 'update_profile.html', {'form': form})
    except Exception as e:
        logger.error(f"Error updating profile for user {request.user.username}: {str(e)}")
        messages.error(request, _("Erreur lors de la mise à jour du profil"))
        return redirect('dashboard')


@login_required
def dashboard(request):
    if get_student_for_user(request.user):
        return redirect('student_dashboard')

    filieres = Filiere.objects.annotate(
        student_count=Count('etudiants', distinct=True),
        present_count=Count(
            'etudiants__presences',
            distinct=True,
            filter=Q(
                etudiants__presences__date=now().date(),
                etudiants__presences__present=True,
            ),
        ),
    ).order_by('nom')
    filiere_form = FiliereForm(request.POST or None)

    if request.method == 'POST':
        if filiere_form.is_valid():
            filiere_form.save()
            messages.success(request, _("La classe a ete ajoutee avec succes."))
            return redirect('dashboard')
        messages.error(request, _("Impossible d'ajouter la classe. Verifiez le formulaire."))

    stats = {
        'filiere_count': filieres.count(),
        'student_count': Etudiant.objects.count(),
        'attendance_count_today': Presence.objects.filter(date=now().date()).count(),
        'pending_justifications_count': AbsenceJustification.objects.filter(status=AbsenceJustification.STATUS_PENDING).count(),
    }
    return render(
        request,
        'home.html',
        {
            'filieres': filieres,
            'filiere_form': filiere_form,
            'stats': stats,
            'today': now().date(),
        },
    )


@login_required
def select_filiere(request):
    return redirect('dashboard')


@login_required
def class_detail(request, filiere_id):
    filiere = get_object_or_404(Filiere, id=filiere_id)
    students = filiere.etudiants.all().order_by('nom')
    student_form = EtudiantForm(request.POST or None)
    today = now().date()
    attendance_map = {
        presence.etudiant_id: presence.present
        for presence in Presence.objects.filter(etudiant__filiere=filiere, date=today)
    }

    if request.method == 'POST' and request.POST.get('action') == 'add_student':
        if student_form.is_valid():
            student = student_form.save(commit=False)
            student.filiere = filiere
            if student_form.cleaned_data.get('create_student_account'):
                student_user = User.objects.create_user(
                    username=student_form.cleaned_data['student_username'],
                    password=student_form.cleaned_data['student_password'],
                    email=student.email,
                )
                student.user = student_user
            student.save()
            messages.success(request, _("L'etudiant a ete ajoute a la classe."))
            return redirect('class_detail', filiere_id=filiere.id)
        messages.error(request, _("Le formulaire etudiant contient des erreurs."))

    justifications = AbsenceJustification.objects.filter(
        presence__etudiant__filiere=filiere
    ).select_related('presence', 'presence__etudiant')[:8]

    return render(
        request,
        'list_students.html',
        {
            'filiere': filiere,
            'students': students,
            'student_form': student_form,
            'attendance_map': attendance_map,
            'today': today,
            'justifications': justifications,
        },
    )


@login_required
def import_students_excel(request, filiere_id):
    """Importe une liste d'etudiants depuis un fichier Excel"""
    filiere = get_object_or_404(Filiere, id=filiere_id)
    
    # Verifier que ce n'est pas un etudiant qui essaie d'acceder
    if get_student_for_user(request.user):
        messages.error(request, _("Acces non autorise."))
        return redirect('student_dashboard')
    
    form = ImportExcelForm(request.POST or None, request.FILES or None)
    
    if request.method == 'POST' and form.is_valid():
        if not OPENPYXL_AVAILABLE:
            messages.error(request, _("openpyxl n'est pas installe. Installez-le avec: pip install openpyxl"))
            return render(request, 'import_students.html', {'form': form, 'filiere': filiere})
        
        excel_file = request.FILES['excel_file']
        
        try:
            # Charger le fichier Excel
            from openpyxl import load_workbook
            workbook = load_workbook(excel_file)
            worksheet = workbook.active
            
            # Extraire les donnees et creer les etudiants
            imported_count = 0
            error_count = 0
            errors = []
            
            # Supposer que les colonnes sont: Nom (A), Numero Etudiant (B), Email (C)
            # La premiere ligne est l'entete
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Extraire les donnees de la ligne
                    nom = row[0] if len(row) > 0 else None
                    numero_etudiant = row[1] if len(row) > 1 else None
                    email = row[2] if len(row) > 2 else None
                    
                    # Validation des donnees obligatoires
                    if not nom:
                        errors.append(_("Ligne {row_idx}: Le nom est obligatoire").format(row_idx=row_idx))
                        error_count += 1
                        continue
                    
                    # Nettoyer les donnees
                    nom = str(nom).strip()
                    numero_etudiant = str(numero_etudiant).strip() if numero_etudiant else None
                    email = str(email).strip() if email else None
                    
                    # Valider que le nom n'est pas vide apres nettoyage
                    if not nom or nom.lower() == 'none':
                        errors.append(_("Ligne {row_idx}: Le nom ne peut pas etre vide").format(row_idx=row_idx))
                        error_count += 1
                        continue
                    
                    # Verifier si l'etudiant existe deja (par numero ou email)
                    existing_student = None
                    if numero_etudiant and numero_etudiant.lower() != 'none':
                        existing_student = Etudiant.objects.filter(
                            numero_etudiant=numero_etudiant,
                            filiere=filiere
                        ).first()
                    
                    if existing_student:
                        # Mettre a jour l'etudiant existant
                        existing_student.nom = nom
                        if email and email.lower() != 'none':
                            existing_student.email = email
                        existing_student.save()
                        imported_count += 1
                    else:
                        # Creer un nouvel etudiant
                        student = Etudiant(
                            nom=nom,
                            numero_etudiant=numero_etudiant if numero_etudiant and numero_etudiant.lower() != 'none' else None,
                            email=email if email and email.lower() != 'none' else '',
                            filiere=filiere
                        )
                        student.save()
                        imported_count += 1
                
                except Exception as e:
                    error_count += 1
                    errors.append(f"Ligne {row_idx}: {str(e)}")
            
            # Afficher les resultats
            message_text = _("Import termine: {imported_count} etudiant(s) ajoute(s) ou mise a jour.").format(imported_count=imported_count)
            if error_count > 0:
                message_text += " " + _("{error_count} erreur(s) rencontree(s).").format(error_count=error_count)
                for error in errors[:5]:  # Afficher les 5 premieres erreurs
                    messages.warning(request, error)
            
            messages.success(request, message_text)
            return redirect('class_detail', filiere_id=filiere.id)
        
        except Exception as e:
            messages.error(request, _("Erreur lors de la lecture du fichier: {error}").format(error=str(e)))
    
    return render(
        request,
        'import_students.html',
        {
            'form': form,
            'filiere': filiere,
        },
    )


@login_required
def mark_attendance(request, filiere_id):
    return redirect('class_detail', filiere_id=filiere_id)


def _update_student_totals(student):
    student.total_presences = student.presences.filter(present=True).count()
    student.total_absences = student.presences.filter(present=False).count()
    student.save(update_fields=['total_presences', 'total_absences'])


@login_required
@require_POST
def save_attendance(request, filiere_id):
    filiere = get_object_or_404(Filiere, id=filiere_id)
    students = filiere.etudiants.all()

    for student in students:
        is_present = f'present_{student.id}' in request.POST
        presence, created = Presence.objects.get_or_create(
            etudiant=student,
            date=now().date(),
            defaults={'present': is_present},
        )

        if not created and presence.present != is_present:
            presence.present = is_present
            presence.save(update_fields=['present'])

        _update_student_totals(student)

    messages.success(request, _("L'appel du jour a ete enregistre."))
    return generate_daily_pdf(filiere)


@login_required
def monthly_report(request):
    if get_student_for_user(request.user):
        return redirect('student_dashboard')

    filieres = Filiere.objects.all().order_by('nom')
    report = []
    current_month = now().month
    current_year = now().year

    for filiere in filieres:
        data = []
        for etudiant in filiere.etudiants.all().order_by('nom'):
            total_presences = Presence.objects.filter(
                etudiant=etudiant,
                present=True,
                date__month=current_month,
                date__year=current_year,
            ).count()
            total_absences = Presence.objects.filter(
                etudiant=etudiant,
                present=False,
                date__month=current_month,
                date__year=current_year,
            ).count()
            data.append(
                {
                    'etudiant': etudiant,
                    'total_presences': total_presences,
                    'total_absences': total_absences,
                }
            )
        report.append({'filiere': filiere, 'data': data})

    return render(request, 'monthly_report.html', {'report': report, 'today': now().date()})


@login_required
def generate_monthly_pdf(request, filiere_id):
    filiere = get_object_or_404(Filiere, id=filiere_id)

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer)
    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(180, 800, _("Rapport mensuel - {nom}").format(nom=filiere.nom))
    pdf.setFont("Helvetica", 12)
    pdf.drawString(50, 775, _("Periode : {date}").format(date=now().strftime('%m/%Y')))
    pdf.drawString(50, 750, _("Etudiant"))
    pdf.drawString(300, 750, _("Presences"))
    pdf.drawString(430, 750, _("Absences"))

    y = 725
    for etudiant in filiere.etudiants.all().order_by('nom'):
        total_presences = Presence.objects.filter(
            etudiant=etudiant,
            present=True,
            date__month=now().month,
            date__year=now().year,
        ).count()
        total_absences = Presence.objects.filter(
            etudiant=etudiant,
            present=False,
            date__month=now().month,
            date__year=now().year,
        ).count()

        pdf.drawString(50, y, etudiant.nom)
        pdf.drawString(320, y, str(total_presences))
        pdf.drawString(445, y, str(total_absences))
        y -= 20
        if y < 60:
            pdf.showPage()
            pdf.setFont("Helvetica", 12)
            y = 800

    pdf.save()
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="rapport_{filiere.nom}.pdf"'
    return response


def generate_daily_pdf(filiere):
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer)
    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(180, 800, _("Appel du jour - {nom}").format(nom=filiere.nom))
    pdf.setFont("Helvetica", 12)
    pdf.drawString(50, 775, _("Date : {date}").format(date=now().strftime('%d/%m/%Y')))
    pdf.drawString(50, 750, _("Etudiant"))
    pdf.drawString(320, 750, _("Presence"))

    y = 725
    for student in filiere.etudiants.all().order_by('nom'):
        presence = Presence.objects.filter(etudiant=student, date=now().date()).first()
        status = _("Present") if presence and presence.present else _("Absent")
        pdf.drawString(50, y, student.nom)
        pdf.drawString(320, y, status)
        y -= 20
        if y < 60:
            pdf.showPage()
            pdf.setFont("Helvetica", 12)
            y = 800

    pdf.save()
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="appel_{filiere.nom}_{now().date()}.pdf"'
    return response


@login_required
def student_dashboard(request):
    student = get_student_for_user(request.user)
    if not student:
        return redirect('dashboard')

    presences = student.presences.select_related('justification').all()
    absences = [presence for presence in presences if not presence.present]
    justified_count = sum(1 for presence in absences if hasattr(presence, 'justification'))

    # Get class schedule for the student's filiere
    schedule = ClassSchedule.objects.filter(filiere=student.filiere).order_by('day_of_week', 'start_time')

    stats = {
        'class_name': student.filiere.nom,
        'presence_count': student.total_presences,
        'absence_count': student.total_absences,
        'justified_count': justified_count,
        'pending_absence_count': student.total_absences - justified_count,
    }
    return render(
        request,
        'student_dashboard.html',
        {
            'student': student,
            'presences': presences,
            'stats': stats,
            'schedule': schedule,
        },
    )


@login_required
def justify_absence(request, presence_id):
    student = get_student_for_user(request.user)
    if not student:
        return redirect('dashboard')

    presence = get_object_or_404(Presence, id=presence_id, etudiant=student, present=False)
    justification = getattr(presence, 'justification', None)
    form = AbsenceJustificationForm(request.POST or None, instance=justification)

    if request.method == 'POST' and form.is_valid():
        justification = form.save(commit=False)
        justification.presence = presence
        justification.status = AbsenceJustification.STATUS_PENDING
        justification.save()
        messages.success(request, _("Votre justification a bien ete envoyee."))
        return redirect('student_dashboard')

    return render(
        request,
        'justify_absence.html',
        {
            'student': student,
            'presence': presence,
            'form': form,
        },
    )


@login_required
def manage_justifications(request):
    """
    Permet aux professeurs de gerer les justifications d'absence (approuver/rejeter).
    """
    if get_student_for_user(request.user):
        messages.error(request, _("Acces non autorise."))
        return redirect('student_dashboard')

    pending_justifications = AbsenceJustification.objects.filter(
        status=AbsenceJustification.STATUS_PENDING
    ).select_related('presence__etudiant', 'presence__etudiant__filiere').order_by('-submitted_at')

    if request.method == 'POST':
        justification_id = request.POST.get('justification_id')
        action = request.POST.get('action')  # 'approve' or 'reject'

        justification = get_object_or_404(AbsenceJustification, id=justification_id)

        if action == 'approve':
            justification.status = AbsenceJustification.STATUS_APPROVED
            messages.success(request, _("Justification de {nom} approuvee.").format(nom=justification.presence.etudiant.nom))
        elif action == 'reject':
            justification.status = AbsenceJustification.STATUS_REJECTED
            messages.warning(request, _("Justification de {nom} rejetee.").format(nom=justification.presence.etudiant.nom))
        else:
            messages.error(request, _("Action invalide."))
            return redirect('manage_justifications')

        justification.reviewed_at = now()
        justification.save()
        return redirect('manage_justifications')

    return render(request, 'manage_justifications.html', {
        'pending_justifications': pending_justifications,
    })


# ======================== NEW SIDEBAR VIEWS ========================

@login_required
def general_information(request):
    return render(request, 'general_information.html')

@login_required
def calendar_view(request):
    student = get_student_for_user(request.user)
    current_dt = localtime()
    current_day_name = current_dt.strftime('%A')
    current_time = current_dt.time()

    if student:
        filieres = [student.filiere]
    else:
        filieres = list(Filiere.objects.prefetch_related('schedules').order_by('nom'))

    calendar_sections = []
    for filiere in filieres:
        schedules = list(filiere.schedules.all())
        if student:
            schedules = list(
                ClassSchedule.objects.filter(filiere=filiere)
            )
        calendar_sections.append(
            build_schedule_section(filiere, schedules, current_day_name, current_time)
        )

    has_any_schedule = any(section['total_sessions'] for section in calendar_sections)

    return render(
        request,
        'calendar.html',
        {
            'student': student,
            'calendar_sections': calendar_sections,
            'has_any_schedule': has_any_schedule,
        },
    )

@login_required
def submitted_documents(request):
    return render(request, 'submitted_documents.html')

@login_required
def academic_tracking(request):
    return render(request, 'academic_tracking.html')

@login_required
def academic_grades(request):
    return render(request, 'academic_grades.html')

@login_required
def absences_classes(request):
    student = get_student_for_user(request.user)
    if not student:
        return redirect('dashboard')
    presences = student.presences.select_related('justification').filter(present=False).order_by('-date')
    justified_count = sum(1 for p in presences if hasattr(p, 'justification'))
    stats = {
        'total': presences.count(),
        'justified': justified_count,
        'unjustified': presences.count() - justified_count,
    }
    return render(request, 'absences_classes.html', {'student': student, 'presences': presences, 'stats': stats})

@login_required
def absences_tests(request):
    return render(request, 'absences_tests.html')

@login_required
def absences_exams(request):
    return render(request, 'absences_exams.html')

@login_required
def cheating(request):
    return render(request, 'cheating.html')

@login_required
def sent_sms(request):
    return render(request, 'sent_sms.html')
